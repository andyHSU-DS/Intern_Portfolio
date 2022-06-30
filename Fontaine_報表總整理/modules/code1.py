import numpy as np 
import pandas as pd 
import functools
from openpyxl import load_workbook
from pandas import ExcelWriter 
from openpyxl import Workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


def compare_list(l1,l2):
    for x in l1:
        for j in l2:
            if(x==j):
                return 1
#-----------------------讀取所有的案件號--------------------------------------------
def read_all_case_number(df):
    full_case_number=df["案件號"].tolist()
    full_case_number=list(set(full_case_number))
    return full_case_number
#-----------------------讀取所有的案件號相對應的名子並做成一個 dataframe--------------------------------------------
def name_dataframe(full_case_number,data):
    name_dataframe=pd.DataFrame(columns=np.arange(0,len(full_case_number)+1))
    name_dataframe=name_dataframe.transpose()
    for i in range(0,len(full_case_number)):
        case_number=full_case_number[i]
        temp_1_list=data[data["案件號"]==case_number]["戶名"].tolist()
        temp_1=list(set(temp_1_list))
        name_dataframe[str(case_number)]=pd.Series(temp_1)
    name_dataframe=name_dataframe.transpose()
    col=np.arange(1,name_dataframe.shape[1]+1)
    name_dataframe.columns=col
    name_dataframe=name_dataframe.transpose()
    return name_dataframe
#-----------------------找出dataframe裡不同案件號有相同名子的案件號--------------------------------------------
def get_case_number_result(name_dataframe):
    first_result_number=pd.DataFrame(columns=np.arange(0,name_dataframe.shape[1]+1))
    first_result_number=first_result_number.transpose()
    for i in range(1,name_dataframe.shape[1]+1):
        n1=name_dataframe[str(i)].tolist()
        temp_result_number=[]
        for j in range(1,name_dataframe.shape[1]+1):
            n2=name_dataframe[str(j)].tolist()
            if (compare_list(n1,n2)==1):
                temp_result_number.append(j)
        result_number=list(set(temp_result_number))
        first_result_number[str(i)]=pd.Series(result_number)
    first_result_number=first_result_number.transpose()
    col=np.arange(0,first_result_number.shape[1])
    first_result_number.columns=col
    return first_result_number

def get_second_result_number(first_result_number,name_df):
    second=pd.DataFrame(columns=np.arange(0,name_df.shape[1]+1))
    second=second.transpose()
    for i in range(0,first_result_number.shape[0]):
        l1=first_result_number.iloc[i].tolist()
        temp=[]
        for j in range(0,name_df.shape[1]):
            l2=first_result_number.iloc[j].tolist()
            if(compare_list(l1,l2)==1):
                temp.append(j+1)
        result_number=list(set(temp))
        second[str(i)]=pd.Series(result_number)
    second=second.transpose()
    col=np.arange(0,second.shape[1])
    second.columns=col
    return second

def get_final_data(df,name_df):
    c=get_second_result_number(df,name_df)
    for _ in range(30):
        c=get_second_result_number(c,name_df)
    c=c.drop_duplicates(subset=None, keep='first', inplace=False)
    return c

def real_final_df(final_data):
    categroy=final_data.shape[0]
    temp=np.arange(1,categroy+1)
    final_data["歸戶編號"]=temp
    return final_data

def last_function_for_data(df,data):
    data["temp_歸戶編號"]= ""
    for i in range(df.shape[0]):
        final_case_number=df.iloc[i].tolist()
        final_case_number=final_case_number[:-1]
        cn=df["歸戶編號"][i]
        for k in range(len(final_case_number)):
            f_number=final_case_number[k]
            for j in range(data.shape[0]):
                if(data["案件號"][j]==f_number):
                    data["temp_歸戶編號"][j]=cn
    #這是因為如果前面做過了，又insert就會出錯
    if '歸戶編號' not in data.columns:
        data.insert(loc=0,column="歸戶編號",value=data["temp_歸戶編號"])
        data=data.drop(["temp_歸戶編號"],axis=1)
    else:
        pass
    return data

def main_fnc(data):
    full_case_number = read_all_case_number(data)
    name_df=name_dataframe(full_case_number,data)
    fir=get_case_number_result(name_df)
    df=get_final_data(fir,name_df)
    df=real_final_df(df)
    final=last_function_for_data(df,data)
    return final
    
def write_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.append(list(df.columns))
    for i in range(df.shape[0]):
        row=df.iloc[i].values
        row = list(row)
        ws.append(row)
    row=df.shape[0]+1
    column=df.shape[1]
    column=chr(64+column)
    tab = Table(displayName="Table",ref="A1:"+str(column)+str(row))
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(showFirstColumn=True,
                         showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    #---------------------------------------設定邊框--------------------------------------------
    rows = ws.max_row
    cols = ws.max_column
    font2 = Font(name='微软雅黑', size=11)
    line_t = Side(style='thin', color='000000') 
    line_m = Side(style='medium', color='000000')
    alignment = Alignment(horizontal='center', vertical='center')

    border1 = Border(top=line_m, bottom=line_t, left=line_t, right=line_t)
    border2 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

    sty1 = NamedStyle(name='sty1', font=font2, 
              border=border1, alignment=alignment)
    sty2 = NamedStyle(name='sty2', font=font2, border=border2, alignment=alignment)
    for r in range(1, rows+1):
        for c in range(1, cols+1):
            if r == 2:
                ws.cell(r, c).style = sty1
            else:
                ws.cell(r, c).style = sty2
        #-------------------------------------反白------------------------------------------------
    return wb 