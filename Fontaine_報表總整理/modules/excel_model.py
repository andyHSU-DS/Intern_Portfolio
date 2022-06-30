import numpy as np 
import pandas as pd 
from pandas import ExcelWriter 
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

class Excel_Work():
    def write_excel(self,df,line,tabel_style,tabel_style_name):       
        wb = Workbook()
        ws = wb.active
        ws.append(list(df.columns))
        for i in range(df.shape[0]):
            row=df.iloc[i].values
            row = list(row)
            ws.append(row)
        row=df.shape[0]+1
        column=df.shape[1]
        column=chr(64+column)
        
        tab = Table(displayName="Table",ref="A1:"+str(column)+str(row))
        # Add a default style with striped rows and banded columns
        if tabel_style == True:
            style = TableStyleInfo(name=str(tabel_style_name), showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        else:
            style = TableStyleInfo(showFirstColumn=True,showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        #---------------------------------------設定邊框--------------------------------------------
        if line == True :
            rows = ws.max_row
            cols = ws.max_column
            font2 = Font(name='微软雅黑', size=11)
            line_t = Side(style='thin', color='000000') 
            line_m = Side(style='medium', color='000000')
            alignment = Alignment(horizontal='center', vertical='center')

            border1 = Border(top=line_m, bottom=line_t, left=line_t, right=line_t)
            border2 = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

            sty1 = NamedStyle(name='sty1', font=font2, 
                    border=border1, alignment=alignment)
            sty2 = NamedStyle(name='sty2', font=font2, border=border2, alignment=alignment)

            for r in range(1, rows+1):
                for c in range(1, cols+1):
                    if r == 2:
                        ws.cell(r, c).style = sty1
                    else:
                        ws.cell(r, c).style = sty2
            #-------------------------------------反白------------------------------------------------
        return wb
    def append_excel(self,excel_path,df,excel_sheet_name):
        book = load_workbook(excel_path)
        writer = ExcelWriter(excel_path, engine='openpyxl')  # pylint: disable=abstract-class-instantiated
        writer.book = book
        df.to_excel(writer,str(excel_sheet_name),index=False)
        writer.save()
    def pandas_method(self,df,output_path):
        df.to_excel(output_path)



check = pd.DataFrame({
    "a":[1,2,3,4,5]
})


excel = Excel_Work()
wb=excel.write_excel(df=check,line=True,tabel_style=False,tabel_style_name="TableStyleMedium9")
wb.save(r"D:\My Documents\kenc\桌面\testing.xlsx")

# excel.append_excel(excel_path=r"D:\My Documents\kenc\桌面\testing.xlsx",df=check,excel_sheet_name="testing")
